import numpy as np
#from openpyxl.workbook import Workbook
import pandas as pd
from sqlalchemy import create_engine
from datetime import timedelta
from dateutil.relativedelta import relativedelta
from pprint import pprint
import os
import openpyxl
import random

print('KPI Framework started')

outname = 'KPIFramework_Python.csv'
#outdir = 'C:/Users/nickh/OneDrive/Documents/Projects/cookkpi/dashboard/src/assets/Attributes/dashboard_data'

current_dir = os.path.dirname(os.path.realpath(__file__))
three_levels_back = os.path.abspath(os.path.join(current_dir, "../../..")).replace('\\', '/')
cookpi_per_pi = 'assets/Attributes/dashboard_data/cookpi_per_pi.xlsx'
d_date = 'assets/Attributes/dashboard_data/d_date.csv'
outdir='assets/Attributes/dashboard_data'


if not os.path.exists(outdir):
    os.mkdir(outdir)

fullname = os.path.join(outdir, outname)

#f_kpi_synthetix = pd.DataFrame(pd.read_csv(
#    "assets/Attributes/f_kpi_synthetix.csv", sep=',', decimal='.', index_col=False))

f_kpi_defilama = pd.DataFrame(pd.read_csv(
    "assets/Attributes/blockchain_api_data/DefiLamaData/f_kpi_DefiLama.csv", sep=',', decimal='.', index_col=False))

f_kpi_coingecko = pd.DataFrame(pd.read_csv(
    "assets/Attributes/blockchain_api_data/CoinGeckoData/f_kpi_coingecko.csv", sep=',', decimal='.', index_col=False))



facts = [f_kpi_defilama,f_kpi_coingecko]#f_kpi_synthetix,
f_kpi = pd.concat(facts)

tmp_d_kpi = pd.read_excel(open(cookpi_per_pi, 'rb'),
              sheet_name='d_kpi')
tmp_d_kpi2 = tmp_d_kpi[(tmp_d_kpi.live == 1)]
d_kpi = tmp_d_kpi2[['d_kpi_id','KPIName','KPIGroup','Calculation','numerator_id',
                   'denominator_id','AggregateNum','AggregateDenom','KPIType','Sorting']]

#d_kpi['IsCalculatedKPI'] = np.where(d_kpi['denominator_id']!='0', '1', '0')

d_date = pd.DataFrame(pd.read_csv(
    d_date, sep=',', index_col=False))

# Dataframes Dimensions/attributes

keys = ['d_kpi_id', 'd_date_id']  # ,'d_level0_id','d_level1_id','d_level2_id'
Grain = ['full_date', 'LD_Month', 'LD_Quarter', 'LD_Year']
GrainName = ['D', 'M', 'Q', 'Y']
KPIIDList = d_kpi['d_kpi_id']
AggregateNumList = d_kpi['AggregateNum']
AggregateDenomList = d_kpi['AggregateDenom']
KPITypeList = d_kpi['KPIType']


def AggregateNumerator(Aggregate):
    if Aggregate == 1:
        CalculationString = "'sum'"
        return CalculationString
    elif Aggregate == 2:
        CalculationString = "'mean'"
        return CalculationString
    elif Aggregate == 3:
        CalculationString = "'max'"
        return CalculationString


def AggregateDenominator(Aggregate):
    if Aggregate == 1:
        CalculationString = "'sum'"
        return CalculationString
    elif Aggregate == 2:
        CalculationString = "'mean'"
        return CalculationString
    elif Aggregate == 3:
        CalculationString = "'max'"
        return CalculationString


"""
-Create all newly created kpi's by searching for the numerator and the denominator in the existing kpi's
- Concat the old f_kpi with the newly created kpi's
"""

d_kpi_calculate = d_kpi[['d_kpi_id', 'numerator_id', 'denominator_id']]
a = d_kpi_calculate[d_kpi_calculate.denominator_id != 0]
print(d_kpi)
Denomvaluedf = f_kpi[['d_kpi_id', 'd_level0_id',
                      'd_level1_id', 'd_level2_id', 'd_date_id', 'Numerator']]
f_kpi_cook = pd.merge(a, Denomvaluedf, how='left',
                      left_on='denominator_id', right_on='d_kpi_id')
f_kpi_cook.rename(columns={"Numerator": "Denominator",
                  "d_kpi_id_x": "d_kpi_id"}, inplace=True)
print(f_kpi_cook)
f_kpi_cook = f_kpi_cook.drop(
    columns=['d_kpi_id_y', 'numerator_id', 'denominator_id'], axis=1)
f_kpi_cook["Join"] = f_kpi_cook['d_kpi_id'].astype(str) + f_kpi_cook["d_level0_id"].astype(str) + f_kpi_cook["d_level1_id"].astype(
    str) + f_kpi_cook["d_level2_id"].astype(str) + f_kpi_cook["d_date_id"].astype(str)  # .astype(str)
print(f_kpi_cook)
f_kpi_cook2 = pd.merge(a, Denomvaluedf, how='left',
                       left_on='numerator_id', right_on='d_kpi_id')
f_kpi_cook2.rename(columns={"d_kpi_id_x": "d_kpi_id"}, inplace=True)
f_kpi_cook2 = f_kpi_cook2.drop(
    columns=['d_kpi_id_y', 'numerator_id', 'denominator_id'], axis=1)
f_kpi_cook2["Join"] = f_kpi_cook2['d_kpi_id'].astype(str) + f_kpi_cook2["d_level0_id"].astype(str) + f_kpi_cook2["d_level1_id"].astype(
    str) + f_kpi_cook2["d_level2_id"].astype(str) + f_kpi_cook2["d_date_id"].astype(str)  # .astype(str)


f_kpi_cook3 = f_kpi_cook.merge(f_kpi_cook2, on='Join', how='left')
f_kpi_cook3.drop(f_kpi_cook3.filter(regex='_y$').columns, axis=1, inplace=True)
f_kpi_cook3.columns = f_kpi_cook3.columns.str.rstrip('_x')
f_kpi_cook3 = f_kpi_cook3.drop(columns=['Join'], axis=1).fillna(0)


f_kpi_cook4= f_kpi_cook3.dropna(subset=['Numerator','Denominator'])
f_kpi_cook4.reset_index()

f_kpi_cook4.to_csv(
    r'assets/Attributes/dashboard_data/KPIFramework_Pythontmp4.csv', index=False)

f_kpi = pd.concat([f_kpi, f_kpi_cook4])
f_kpi.reset_index()

f_kpi['d_kpi_id'] = f_kpi['d_kpi_id'].astype('int64')
f_kpi['d_level0_id'] = f_kpi['d_level0_id'].astype('int64')
f_kpi['d_level1_id'] = f_kpi['d_level1_id'].astype('int64')
f_kpi['d_level2_id'] = f_kpi['d_level2_id'].astype('int64')
f_kpi['d_date_id'] = f_kpi['d_date_id'].astype('int64')


f_kpi['fact_id'] = f_kpi['d_kpi_id'].astype(str)+f_kpi['d_level0_id'].astype(str)+f_kpi['d_level1_id'].astype(str)+f_kpi['d_level2_id'].astype(str)
f_kpi.reset_index()


determinelastdates = f_kpi[['fact_id','d_date_id']]

determinelastdates = pd.merge(determinelastdates,d_date, how='left',
                      left_on='d_date_id', right_on='d_date_id')
determinelastdates = determinelastdates.reset_index()            
agglist = ['fact_id','int_month']
determinelastdatesMonth = determinelastdates.groupby(['fact_id', 'int_month'])['d_date_id'].max().reset_index()
determinelastdatesQuarter = determinelastdates.groupby(['fact_id', 'int_quarter'])['d_date_id'].max().reset_index()
determinelastdatesYear = determinelastdates.groupby(['fact_id', 'int_year'])['d_date_id'].max().reset_index()
determinelastdatesMonth.drop('int_month', axis=1, inplace=True)
determinelastdatesQuarter.drop('int_quarter', axis=1, inplace=True)
determinelastdatesYear.drop('int_year', axis=1, inplace=True)


df_list = [f_kpi, d_kpi, d_date]
"""
-Join all reference tables to the dataframe with a loop
-Group by all relevant columns to calculate the numerator and denominator
-Loop through to repeat this for all the grains provided (time-grains)
-Create file for every grain calculated
"""
to_csv_filetmp =[]
to_csv_filetmp.clear()
for z, o in zip(Grain, GrainName):
    agglist = ['d_kpi_id', 'd_level0_id', 'd_level1_id', 'd_level2_id', z, o]
    for d, b, l, c in zip(KPIIDList, KPITypeList ,AggregateNumList, AggregateDenomList):
        dftmp = df_list[0]
        dftmp.round(2)
        df = dftmp[(dftmp.d_kpi_id == d)]
       # df['Numerator'] = df['Numerator'].astype(int)
        for i, x in zip(df_list[1:], range(len(keys))):
            df = df.merge(i, how='left', on=keys[x])
        if b == 'Snapshot':
            if o == 'D':
                dfgroupedtmp = df
                dfgrouped = dfgroupedtmp.groupby(agglist, as_index=False).agg({'Numerator': [eval(
                AggregateNumerator(l))], 'Denominator': [eval(AggregateDenominator(c))]})
            elif o == 'M':
                dfgroupedtmp = df.merge(determinelastdatesMonth)#, , how='inner'on=['fact_id','d_date_id']
                dfgrouped = dfgroupedtmp.groupby(agglist, as_index=False).agg({'Numerator': [eval(
                AggregateNumerator(l))], 'Denominator': [eval(AggregateDenominator(c))]})
            elif o == 'Q':
                dfgroupedtmp = df.merge(determinelastdatesQuarter)#, how='inner', on=['fact_id','d_date_id']
                dfgrouped = dfgroupedtmp.groupby(agglist, as_index=False).agg({'Numerator': [eval(
                AggregateNumerator(l))], 'Denominator': [eval(AggregateDenominator(c))]})
            elif o == 'Y':
                dfgroupedtmp = df.merge(determinelastdatesYear)#, how='inner', on=['fact_id','d_date_id']
                dfgrouped = dfgroupedtmp.groupby(agglist, as_index=False).agg({'Numerator': [eval(
                AggregateNumerator(l))], 'Denominator': [eval(AggregateDenominator(c))]}) 
        else:     
            dfgrouped = df.groupby(agglist, as_index=False).agg({'Numerator': [eval(
                AggregateNumerator(l))], 'Denominator': [eval(AggregateDenominator(c))]})
        grouped_multiple = dfgrouped
        grouped_multiple.columns = grouped_multiple.columns.droplevel(1)
        grouped_multiple['Grain'] = o
        grouped_multiple.rename(columns={z: 'Period_int'}, inplace=True)
        grouped_multiple.rename(columns={o: 'PeriodName'}, inplace=True)
        
        to_csv_filetmp.append(grouped_multiple)

KPIFramework = pd.concat(to_csv_filetmp)      
KPIFramework.to_csv(
            r'assets/Attributes/dashboard_data/KPIFramework_Python.csv', index=False)  #_' + str(o) + '      

"""
KPIFrameworkDay = pd.DataFrame(pd.read_csv(
    r'assets/Attributes/dashboard_data/KPIFramework_Python_D.csv',low_memory=False))
KPIFrameworkMonth = pd.DataFrame(pd.read_csv(
    r'assets/Attributes/dashboard_data/KPIFramework_Python_M.csv',low_memory=False))
KPIFrameworkQuarter = pd.DataFrame(pd.read_csv(
    r'assets/Attributes/dashboard_data/KPIFramework_Python_Q.csv',low_memory=False))
KPIFrameworkYear = pd.DataFrame(pd.read_csv(
    r'assets/Attributes/dashboard_data/KPIFramework_Python_Y.csv',low_memory=False))
"""

"""
-Concatenate the files created in previous step and create a new file with all grains concatenated. Distinguish by filtering on column 'grain'
-Convert period integer to datetime (needed later for visualisation and calculations)
-Calculate an extra column named 'Period in lp' meaning: the data of the previous period. This will be used later to join the numerator and denominator of the previous period
"""
 
KPIFrameworktmp = KPIFramework#pd.concat([KPIFrameworkDay, KPIFrameworkMonth, KPIFrameworkQuarter, KPIFrameworkYear])
KPIFrameworktmp["Period_int"] = pd.to_datetime(KPIFrameworktmp["Period_int"])
KPIFrameworktmp['Period_int_lp'] = KPIFrameworktmp[['Grain', 'Period_int']].apply(lambda x: x['Period_int'] + timedelta(days=+1) if x['Grain'] == 'D' else x['Period_int'] + timedelta(
    days=+365) if x['Grain'] == 'Y' else x['Period_int'] + relativedelta(months=+1) if x['Grain'] == 'M' else x['Period_int'] + relativedelta(months=+3) if x['Grain'] == 'Q' else x['Period_int'], axis=1)

"""
-Create two row-id's. One is for the dataframe itself and one is manipulated to be used to join
-Create second KPIFramework and only select numerator and denominator with the row_id_lp
-Merge the two ID's to add last period numerator and denominator to existing framework
-Drop unnesseccary columns
-Save file
"""

KPIFrameworktmp['Row_id'] = KPIFrameworktmp.apply(lambda x: x['Period_int'].strftime(
    '%Y%m%d') + str(x['d_kpi_id']) + str(x['d_level0_id']) + str(x['d_level1_id']) + str(x['d_level2_id']) + str(x['Grain']), axis=1)
KPIFrameworktmp['Row_id_lp'] = KPIFrameworktmp.apply(lambda x: x['Period_int_lp'].strftime(
    '%Y%m%d') + str(x['d_kpi_id']) + str(x['d_level0_id']) + str(x['d_level1_id']) + str(x['d_level2_id']) + str(x['Grain']), axis=1)

KPIFramework_LP = KPIFrameworktmp[['Row_id_lp', 'Numerator', 'Denominator']]
KPIFrameworktmp = KPIFrameworktmp.merge(KPIFramework_LP, how='left', left_on=[
                                        'Row_id'], right_on=['Row_id_lp'], suffixes=("", "_LP"))

KPIFrameworktmp2 = KPIFrameworktmp.drop(
    columns=['Row_id_lp', 'Row_id_lp_LP', 'Row_id'], axis=1)#, 'Unnamed: 0'
KPIFrameworktmp2["Period_int"] = pd.to_datetime(KPIFrameworktmp2["Period_int"], format="%Y-%m-%d")

KPIFrameworktmp2.to_csv(
    r'assets/Attributes/dashboard_data/KPIFramework_Pythontmp.csv', index=False)

#df = pd.read_csv(r'assets/Attributes/dashboard_data/KPIFramework_Python.csv')

# Drop all rows
#df.drop(df.index, inplace=True)
#print(df)
#df.to_csv(fullname, index=False)

KPIFrameworkOld = pd.DataFrame(pd.read_csv(
    r'assets/Attributes/dashboard_data/KPIFramework_Python.csv',low_memory=False))
KPIFrameworkOld["Period_int"] = pd.to_datetime(KPIFrameworkOld["Period_int"], format="%Y-%m-%d")
KPIFrameworkNew = pd.concat(
    [KPIFrameworkOld, KPIFrameworktmp2]).drop_duplicates().reset_index(drop=True)

# KPIFrameworkNew = KPIFrameworktmp  #Zet deze aan als je helemaal een nieuw framework wilt bouwen
#KPIFrameworkNew2 = KPIFrameworkNew.drop_duplicates(
#    subset=['d_kpi_id', 'd_level0_id', 'd_level1_id', 'd_level2_id', 'PeriodName', 'Grain'], keep='last').reset_index(drop=True)

#print(KPIFrameworkNew2.columns)

KPIFrameworkNew.to_csv(fullname, index=False)


cookpi_attributestmp = pd.read_excel(open(cookpi_per_pi, 'rb'),
              sheet_name='linktable')
print(cookpi_attributestmp)
kpilevelcount = cookpi_attributestmp.groupby(['d_kpi_id'])['d_kpi_id'].count().reset_index(name="kpilevelcount")
d_kpi = d_kpi.merge(kpilevelcount)

d_kpi.sort_values(by=['Sorting'])

wb = openpyxl.load_workbook(r'assets/Attributes/dashboard_data/cookpi_per_pi.xlsx')

sheetcount = len(wb.sheetnames)
sheetnames = wb.sheetnames
sheetnames.remove('d_kpi')
sheetnames.remove('linktable')

def AggregateNumDenom(Calculation):
    if Calculation == 1:
        CalculationString = "'sum'"
        return CalculationString
    elif Calculation == 2:
        CalculationString = "'sum'"
        return CalculationString
    elif Calculation == 3:
        CalculationString = "'max'"
        return CalculationString
    
KPINumAgg = dict(d_kpi.set_index('KPIName')['AggregateNum'].to_dict())
KPIDenomAgg = dict(d_kpi.set_index('KPIName')['AggregateDenom'].to_dict())

KPINumAggid = dict(d_kpi.set_index('d_kpi_id')['AggregateNum'].to_dict())
KPIDenomAggid = dict(d_kpi.set_index('d_kpi_id')['AggregateDenom'].to_dict())
kpilevelcountdict = dict(d_kpi.set_index('d_kpi_id')['kpilevelcount'].to_dict())

KPIFrameworktmp = pd.DataFrame(
        pd.read_csv(r'assets/Attributes/dashboard_data/KPIFramework_Python.csv',sep=',', decimal='.',low_memory=False))
KPIFrameworktmp["Period_int"] = pd.to_datetime(KPIFrameworktmp["Period_int"], format="%Y-%m-%d")

KPIIDList2 =  cookpi_attributestmp['d_kpi_id'].unique()
KPIIDList3 = [value for value in KPIIDList2 if value in KPIIDList.tolist()]
KPIFrameworklist =[]


for i in KPIIDList3:
    KPIFrameworkloop = KPIFrameworktmp[(KPIFrameworktmp.d_kpi_id ==i)]
    cookpi_attributes = cookpi_attributestmp[(cookpi_attributestmp.d_kpi_id == i)]
    sheettmp = cookpi_attributes[(cookpi_attributes.Level_ID_present =="d_level0_id")] 
    sheettmpl1 = cookpi_attributes[(cookpi_attributes.Level_ID_present =="d_level1_id")] 
    sheettmpl2 = cookpi_attributes[(cookpi_attributes.Level_ID_present =="d_level2_id")] 
    rename = dict(sheettmp.set_index('Level_ID')['Level_ID_present'].to_dict()) 
    renamel1 = dict(sheettmpl1.set_index('Level_ID')['Level_ID_present'].to_dict()) 
    renamel2 = dict(sheettmpl2.set_index('Level_ID')['Level_ID_present'].to_dict()) 
    rename_dict = {}
    rename_dict.update(rename)
    rename_dict.update(renamel1)
    rename_dict.update(renamel2)
    KPIFrameworkloop.rename(columns=rename_dict, inplace = True)
    KPIFrameworklist.append(KPIFrameworkloop)
KPIFramework = pd.concat(KPIFrameworklist)

KPIFramework['d_level0_id']=KPIFramework['d_level0_id'].astype(int)
KPIFramework['d_level1_id']=KPIFramework['d_level1_id'].astype(int)
KPIFramework['d_level2_id']=KPIFramework['d_level2_id'].astype(int)

KPIFramework.to_csv(r'assets/Attributes/dashboard_data/KPIFrameworkEnd.csv', index=False)

columnsdf0 = KPIFramework.columns.tolist()

columnsdf0.remove('d_level1_id')
columnsdf0.remove('d_level2_id')
columnsdf0.remove('Numerator')
columnsdf0.remove('Denominator')
columnsdf0.remove('Numerator_LP')
columnsdf0.remove('Denominator_LP')
columnsdf0.remove('Period_int_lp')

columnsdf1 = KPIFramework.columns.tolist()
columnsdf1.remove('d_level2_id')
columnsdf1.remove('Numerator')
columnsdf1.remove('Denominator')
columnsdf1.remove('Numerator_LP')
columnsdf1.remove('Denominator_LP')
columnsdf1.remove('Period_int_lp')

columnsdf2 = KPIFramework.columns.tolist()
columnsdf2.remove('Numerator')
columnsdf2.remove('Denominator')
columnsdf2.remove('Numerator_LP')
columnsdf2.remove('Denominator_LP')
columnsdf2.remove('Period_int_lp')


dfl02 = []
dfl12 = []
dfl22 = []
dflmasterfront =[]
dfl0Compare2 =[]
dfl1Compare2= []
dfl2Compare2= []

dfl02.clear()
dfl12.clear()
dfl22.clear()
dfl0Compare2.clear()
dfl1Compare2.clear()
dfl2Compare2.clear()

def generate_hex():
    return '#' + ''.join(np.random.choice(list('0123456789ABCDEF'), size=6))

cols =['LevelName', 'LevelNameShort', 'LevelDescription',
           'LevelEntitytype', 'LevelColor','Filter1','FilterName','LevelEntitylogo','Filter1Color','InitialShow']
LevelColors= []

for i in KPIIDList3:
    print(f'nowlooping kpi id: {i}')
    KPIFramework_iterate = KPIFramework[(KPIFramework.d_kpi_id ==i)]
    KPIFrameworkl0 = KPIFramework_iterate.groupby(columnsdf0, as_index=False).agg(
    {'Denominator': eval(AggregateNumDenom(KPIDenomAggid[i])), 'Numerator': eval(AggregateNumDenom(KPINumAggid[i])), 'Denominator_LP': eval(AggregateNumDenom(KPIDenomAggid[i])), 'Numerator_LP': eval(AggregateNumDenom(KPINumAggid[i]))})
    
    KPIFrameworkl1 = KPIFramework_iterate.groupby(columnsdf1, as_index=False).agg(
    {'Denominator': eval(AggregateNumDenom(KPIDenomAggid[i])), 'Numerator': eval(AggregateNumDenom(KPINumAggid[i])), 'Denominator_LP': eval(AggregateNumDenom(KPIDenomAggid[i])), 'Numerator_LP': eval(AggregateNumDenom(KPINumAggid[i]))})
    
    KPIFrameworkl2 = KPIFramework_iterate.groupby(columnsdf2, as_index=False).agg(
    {'Denominator': eval(AggregateNumDenom(KPIDenomAggid[i])), 'Numerator': eval(AggregateNumDenom(KPINumAggid[i])), 'Denominator_LP': eval(AggregateNumDenom(KPIDenomAggid[i])), 'Numerator_LP': eval(AggregateNumDenom(KPINumAggid[i]))})
    
    cookpi_attributes = cookpi_attributestmp[(cookpi_attributestmp.d_kpi_id == i)]
    sheettmp = cookpi_attributes[(cookpi_attributes.Level_ID_present =="d_level0_id")] 

    #df.rename(columns={"A": "a", "B": "c"})
    sheet = dict(sheettmp.set_index('Level_ID_present')['Join_ID'].to_dict()) 
    sheetchangeback = dict(sheettmp.set_index('Join_ID')['Level_ID_present'].to_dict()) 
    keys = ['d_kpi_id']
    keyslist = list(sheet.values())
    for j in keyslist:
        keys.append(j)
    d_level0 = pd.read_excel(open(cookpi_per_pi, 'rb'),
              sheet_name=keyslist[0].split('_')[0])
    # Apply the function to 'LevelColor' column
    #d_level0['LevelColor'] = d_level0.apply(lambda row: generate_hex() if row['LevelColor'] == 0 else row['LevelColor'], axis=1)
    LevelColors.append(d_level0[['LevelColor','LevelName']])
    d_level0 = d_level0.fillna(0)
    d_level0 = d_level0.iloc[:, : 9]
    print(d_level0)
    Filter = pd.read_excel(open(cookpi_per_pi,'rb'),
              sheet_name="Filter")
    filtered_df = Filter[(Filter['Active'] == 'y') & (Filter['Filter'] == 'Filter1')]
    filtered_df = filtered_df[['LevelEntitytype', 'FilterName']]

    d_level0 = d_level0.merge(filtered_df, how='left', on='LevelEntitytype')
    no_join_columns = [column for column in d_level0.columns if column.endswith('_NoJoin')]
    d_level0 = d_level0.drop(columns=no_join_columns)
    
    d_level0["InitialShow"].fillna(0, inplace = True)
    d_level0["Filter1"].fillna("Overig", inplace = True)
    d_level0["Filter1Color"].fillna("NotDefined", inplace = True)
    d_level0['InitialShow']=d_level0['InitialShow'].astype(int)
    d_level0 = d_level0.rename(columns={c: c+'_0' for c in d_level0.columns if c in cols})
    d_level0[d_level0.columns[0]]=d_level0[d_level0.columns[0]].astype(np.int64)
    if kpilevelcountdict[i]>1:
        sheettmpl1 = cookpi_attributes[(cookpi_attributes.Level_ID_present =="d_level1_id")]
        sheetl1 = dict(sheettmpl1.set_index('Level_ID_present')['Join_ID'].to_dict())
        sheetchangebackl1 = dict(sheettmpl1.set_index('Join_ID')['Level_ID_present'].to_dict()) 
        keyslistl1 = list(sheetl1.values())
        for t in keyslistl1:
            keys.append(t)
        d_level1 = pd.read_excel(open(cookpi_per_pi, 'rb'),
                  sheet_name=keyslistl1[0].split('_')[0])
        d_level1 = d_level1.iloc[:, : 9]
        no_join_columns = [column for column in d_level1.columns if column.endswith('_NoJoin')]
        d_level1 = d_level1.drop(columns=no_join_columns)
        print(d_level1)
        d_level1["InitialShow"].fillna(0, inplace = True)
        d_level1["Filter1"].fillna("Overig", inplace = True)
        d_level1["Filter1Color"].fillna("NotDefined", inplace = True)
        d_level1['InitialShow']=d_level1['InitialShow'].astype(int)
        d_level1 = d_level1.rename(columns={c: c+'_1' for c in d_level1.columns if c in cols})
        d_level1[d_level1.columns[0]]=d_level1[d_level1.columns[0]].astype(np.int64)
    if kpilevelcountdict[i]>2:
        sheettmpl2 = cookpi_attributes[(cookpi_attributes.Level_ID_present =="d_level2_id")]
        sheetl2 = dict(sheettmpl2.set_index('Level_ID_present')['Join_ID'].to_dict())   

        sheetchangebackl2 = dict(sheettmpl2.set_index('Join_ID')['Level_ID_present'].to_dict()) 
        keyslistl2 = list(sheetl2.values())
        for b in keyslistl2:
            keys.append(b)
        d_level2 = pd.read_excel(open(cookpi_per_pi, 'rb'),
                  sheet_name=keyslistl2[0].split('_')[0])
        d_level2 = d_level2.iloc[:, : 9]
        no_join_columns = [column for column in d_level2.columns if column.endswith('_NoJoin')]
        d_level2 = d_level2.drop(columns=no_join_columns)
        print(d_level2)
        d_level2["InitialShow"].fillna(0, inplace = True)
        d_level2["Filter1"].fillna('Overig', inplace = True)
        d_level2["Filter1Color"].fillna("NotDefined", inplace = True)
        d_level2['InitialShow']=d_level2['InitialShow'].astype(int)
        d_level2 = d_level2.rename(columns={c: c+'_2' for c in d_level2.columns if c in cols})
        d_level2[d_level2.columns[0]]=d_level2[d_level2.columns[0]].astype(np.int64)
    if kpilevelcountdict[i]==1:#'d_level0_id' not in cookpi_attributes['Level_ID_present'].unique().tolist() and 
        KPIFrameworkl0.rename(columns=sheet, inplace = True)  
        df_list_l0 = [KPIFrameworkl0, d_kpi, d_level0]
        dfl0 = df_list_l0[0]
        for i, x in zip(df_list_l0[1:], range(len(keys))):
            dfl0 = dfl0.merge(i, how='left', on=keys[x])#,suffixes=(f'', f'_{x-1}')

        KPIFrameworkl0.rename(columns=sheetchangeback, inplace = True) 
        dfl0.rename(columns=sheetchangeback, inplace = True)  
        dfl0["Period_int"] = pd.to_datetime(dfl0["Period_int"])
        dfl02.append(dfl0.fillna(0))
        dflmasterfront.append(dfl0.fillna(0))
        dfl0Compare2.append(dfl0.fillna(0))
    elif kpilevelcountdict[i]==2: #'d_level1_id' not in cookpi_attributes['Level_ID_present'].unique().tolist() and   
        df_list_l1 = [KPIFrameworkl1, d_kpi, d_level0, d_level1]

        KPIFrameworkl1.rename(columns=sheet, inplace = True)  
        KPIFrameworkl1.rename(columns=sheetl1, inplace = True)  
        dfl1 = df_list_l1[0]
        for i, x in zip(df_list_l1[1:], range(len(keys))):
            print(i)
            print(keys)
            dfl1 = dfl1.merge(i, how='left', on=keys[x])#,suffixes=(f'_{x-2}', f'_{x-1}')

        KPIFrameworkl1.rename(columns=sheetchangeback, inplace = True) 
        KPIFrameworkl1.rename(columns=sheetchangebackl1, inplace = True) 
        dfl1.rename(columns=sheetchangeback, inplace = True) 
        dfl1.rename(columns=sheetchangebackl1, inplace = True) 
        dfl1["Period_int"] = pd.to_datetime(dfl1["Period_int"])
        dfl12.append(dfl1.fillna(0))
        dflmasterfront.append(dfl1.fillna(0))
        dfl1Compare2.append(dfl1.fillna(0))
    elif kpilevelcountdict[i]==3: #'d_level2_id' not in cookpi_attributes['Level_ID_present'].unique().tolist() and 
        df_list_l2 = [KPIFrameworkl2, d_kpi, d_level0, d_level1, d_level2]
        KPIFrameworkl2.rename(columns=sheet, inplace = True)  
        KPIFrameworkl2.rename(columns=sheetl1, inplace = True)  
        KPIFrameworkl2.rename(columns=sheetl2, inplace = True) 
        dfl2 = df_list_l2[0]
        for p, o in zip(df_list_l2[1:], range(len(keys))):
            print(dfl2.columns)
            print(p.columns)
            dfl2 = dfl2.merge(p, how='left', on=keys[o])#,suffixes=(f'', f'_{o-1}')
        KPIFrameworkl2.rename(columns=sheetchangeback, inplace = True) 
        KPIFrameworkl2.rename(columns=sheetchangebackl1, inplace = True) 
        KPIFrameworkl2.rename(columns=sheetchangebackl2, inplace = True) 

        dfl2.rename(columns=sheetchangeback, inplace = True) 
        dfl2.rename(columns=sheetchangebackl1, inplace = True) 
        dfl2.rename(columns=sheetchangebackl2, inplace = True) 
    
        dfl2["Period_int"] = pd.to_datetime(dfl2["Period_int"])
        dfl22.append(dfl2.fillna(0))
        dflmasterfront.append(dfl2.fillna(0))
        dfl2Compare2.append(dfl2.fillna(0))

        keys.clear()

"""
if len(dfl02)!=0:
    dfl0 = pd.concat(dfl02)
    #dfl0['d_level0_id'] = dfl0['d_level0_id'].astype('int64')
    dfl0Compare = pd.concat(dfl0Compare2)
    dfl0.to_csv(r'C:/Users/nickh/OneDrive/Documents/Projects/cookkpi/dashboard/src/assets/Attributes/dashboard_data/dfl0.csv', index=False)
    dfl0.to_parquet(r'C:/Users/nickh/OneDrive/Documents/Projects/cookkpi/dashboard/src/assets/Attributes/dashboard_data/dfl0.parquet', index=False)

if len(dfl12)!=0:
    dfl1 = pd.concat(dfl12)
    dfl1.to_csv(r'C:/Users/nickh/OneDrive/Documents/Projects/cookkpi/dashboard/src/assets/Attributes/dashboard_data/dfl1.csv', index=False)
    dfl1.to_parquet(r'C:/Users/nickh/OneDrive/Documents/Projects/cookkpi/dashboard/src/assets/Attributes/dashboard_data/dfl1.parquet', index=False)
    dfl1Compare = pd.concat(dfl1Compare2)

if len(dfl22)!=0:
    dfl2 = pd.concat(dfl22)
    dfl2.to_csv(r'C:/Users/nickh/OneDrive/Documents/Projects/cookkpi/dashboard/src/assets/Attributes/dashboard_data/dfl2.csv', index=False)
    dfl2.to_parquet(r'C:/Users/nickh/OneDrive/Documents/Projects/cookkpi/dashboard/src/assets/Attributes/dashboard_data/dfl2.parquet', index=False)
    dfl2Compare = pd.concat(dfl2Compare2)

print(dfl0.dtypes)
print(dfl2.dtypes)
"""
dflmasterfront = pd.concat(dflmasterfront)
print(type(dflmasterfront))
print(dflmasterfront.columns)
columns_to_drop = [col for col in dflmasterfront.columns if col.endswith('_y')]
columns_to_drop = [col for col in dflmasterfront.columns if col.endswith('_x')]

# Apply the function to generate hex for rows where value is 0
dflmasterfront = dflmasterfront.drop(columns=columns_to_drop)
dflmasterfront['LevelColor_0'].replace([0], 'Geen', inplace=True)
dflmasterfront['LevelColor_1'].replace([0], 'Geen', inplace=True)
dflmasterfront['LevelColor_2'].replace([0], 'Geen', inplace=True)
dflmasterfront['LevelColor_0']=dflmasterfront['LevelColor_0'].astype(str)
dflmasterfront['LevelColor_1']=dflmasterfront['LevelColor_1'].astype(str)
dflmasterfront['LevelColor_2']=dflmasterfront['LevelColor_2'].astype(str)
dflmasterfront['FilterName_0'].replace([0], 'Onbekend', inplace=True)
dflmasterfront['LevelName_0'].replace([0], 'Geen', inplace=True)
dflmasterfront['LevelName_1'].replace([0], 'Geen', inplace=True)
dflmasterfront['LevelName_2'].replace([0], 'Geen', inplace=True)
dflmasterfront['LevelDescription_0'].replace([0], 'Onbekend', inplace=True)
dflmasterfront['LevelEntitytype_0'].replace([0], 'Onbekend', inplace=True)
dflmasterfront['Filter1_0'].replace([0], 'Geen', inplace=True)
dflmasterfront['LevelDescription_2'].replace([0], 'Onbekend', inplace=True)
dflmasterfront['LevelEntitytype_2'].replace([0], 'Onbekend', inplace=True)
dflmasterfront['Filter1_2'].replace([0], 'Geen', inplace=True)
dflmasterfront.to_csv(r'assets/Attributes/dashboard_data/dflmasterfront.csv', index=False, sep=';')
dflmasterfront.to_parquet(r'assets/Attributes/dashboard_data/dflmasterfront.parquet', index=False)
print('dataframes calculated')
